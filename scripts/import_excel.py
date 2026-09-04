"""
Excel Importer for Wandoor Runners 100k Challenge.
Imports verified historical data from aseem.xlsx (01/09/2026 to 03/09/2026).
Matches athletes with live Strava profiles/avatars.
"""

import os
import json
import re
import openpyxl

def clean_name(name):
    if not name:
        return ""
    text = str(name).strip()
    # Keep printable characters, letters, numbers, spaces, dots, hyphens
    cleaned = re.sub(r"[^\w\s\.\-]", "", text)
    return " ".join(cleaned.split()).strip()


def normalize_for_match(name):
    return re.sub(r"[^a-z0-9]", "", clean_name(name).lower())


def import_from_excel(excel_path="aseem.xlsx"):
    if not os.path.exists(excel_path):
        print(f"[!] Excel file {excel_path} not found.")
        return False

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    sheet = wb["Sheet1"]

    # Load existing Strava snapshot to map real avatars and IDs
    snapshot_file = os.path.join("data", "latest_snapshot.json")
    strava_athletes = []
    if os.path.exists(snapshot_file):
        with open(snapshot_file, "r") as f:
            strava_athletes = json.load(f).get("athletes", [])

    strava_by_id = {str(sa["athlete_id"]): sa for sa in strava_athletes}
    strava_by_norm = {normalize_for_match(sa["name"]): sa for sa in strava_athletes}

    # Explicit known mappings between Excel registrations and Strava profiles
    EXPLICIT_MAP = {
        "nidheeshporur": "40420193",      # Nidheesh Porur (Excel) -> Nidheesh Edappalli (Strava)
        "sarathporur": "61176348",        # Sarath porur (Excel) -> Porur Sarath (Strava)
        "arunnarayanuc": "49239362",      # Arun narayanan UC (Excel) -> Arun Savitha (Strava)
        "arunnarayananuc": "49239362",
        "fahis": "201415432",             # Fahis (Excel) -> Fahis Fahi (Strava)
    }

    dates = ["2026-09-01", "2026-09-02", "2026-09-03"]
    daily_records = {d: [] for d in dates}
    athlete_totals = {}

    target_km = 100.0

    for r in range(7, sheet.max_row + 1):
        sl = sheet.cell(r, 2).value
        raw_name = sheet.cell(r, 3).value
        if not raw_name:
            continue

        name = clean_name(raw_name)
        norm_name = normalize_for_match(name)
        mob = str(sheet.cell(r, 5).value or "").strip()

        # Day 1
        d1 = sheet.cell(r, 6).value
        p1 = str(sheet.cell(r, 7).value or "--").strip()
        t1 = str(sheet.cell(r, 8).value or "").strip()
        d1_km = round(float(d1), 2) if isinstance(d1, (int, float)) else 0.0

        # Day 2
        d2 = sheet.cell(r, 9).value
        p2 = str(sheet.cell(r, 10).value or "--").strip()
        t2 = str(sheet.cell(r, 11).value or "").strip()
        d2_km = round(float(d2), 2) if isinstance(d2, (int, float)) else 0.0

        # Day 3
        d3 = sheet.cell(r, 12).value
        p3 = str(sheet.cell(r, 13).value or "--").strip()
        t3 = str(sheet.cell(r, 14).value or "").strip()
        d3_km = round(float(d3), 2) if isinstance(d3, (int, float)) else 0.0

        # Calculate exact total
        total_km = round(d1_km + d2_km + d3_km, 2)

        # Match with Strava profile if available
        matched_strava = None
        if norm_name in EXPLICIT_MAP:
            matched_strava = strava_by_id.get(EXPLICIT_MAP[norm_name])
        elif norm_name in strava_by_norm:
            matched_strava = strava_by_norm[norm_name]
        else:
            for snorm, sa in strava_by_norm.items():
                if len(norm_name) >= 6 and len(snorm) >= 6 and (norm_name in snorm or snorm in norm_name):
                    matched_strava = sa
                    break

        if matched_strava:
            athlete_id = str(matched_strava["athlete_id"])
            avatar_url = matched_strava.get("avatar_url", "https://d3nn82uaxijpm6.cloudfront.net/sweaters/assets/large.png")
            display_name = clean_name(matched_strava["name"])
        else:
            athlete_id = f"athlete_{sl or norm_name}"
            avatar_url = "https://d3nn82uaxijpm6.cloudfront.net/sweaters/assets/large.png"
            display_name = name

        days_data = [
            ("2026-09-01", d1_km, p1, t1),
            ("2026-09-02", d2_km, p2, t2),
            ("2026-09-03", d3_km, p3, t3)
        ]

        daily_breakdown = {}
        active_days = 0
        best_day = 0.0
        streak = 0
        cum_km = 0.0

        for d_str, dist, pace, time_str in days_data:
            cum_km = round(cum_km + dist, 2)
            daily_breakdown[d_str] = dist

            if dist > 0.0:
                active_days += 1
                streak += 1
                if dist > best_day:
                    best_day = dist
            else:
                streak = 0

            log_entry = {
                "date": d_str,
                "athlete_id": athlete_id,
                "name": display_name,
                "avatar_url": avatar_url,
                "daily_distance_km": dist,
                "daily_runs": 1 if dist > 0 else 0,
                "avg_pace": pace if pace != "--" and pace != "0" else "--",
                "duration": time_str,
                "daily_elev_gain_m": 0,
                "weekly_cumulative_km": cum_km
            }
            daily_records[d_str].append(log_entry)

        pct = min(100.0, round((total_km / target_km) * 100.0, 1))

        athlete_totals[athlete_id] = {
            "athlete_id": athlete_id,
            "sl_no": sl,
            "name": display_name,
            "excel_name": name,
            "avatar_url": avatar_url,
            "mobile": mob,
            "total_challenge_km": total_km,
            "target_km": target_km,
            "active_days": active_days,
            "current_streak": streak,
            "best_day_km": best_day,
            "pct_completed": pct,
            "is_finisher": total_km >= target_km,
            "remaining_km": max(0.0, round(target_km - total_km, 2)),
            "daily_breakdown": daily_breakdown
        }

    # Sort daily records by distance
    for d_str in dates:
        daily_records[d_str].sort(key=lambda x: x["daily_distance_km"], reverse=True)

    # Save to data/club_history.json
    history_file = os.path.join("data", "club_history.json")
    history_payload = {
        "club_name": "Wandoor Runners Monsoon 100k challenge 2026",
        "target_km": target_km,
        "last_updated": "2026-09-04T01:05:00",
        "dates": dates,
        "daily_records": daily_records,
        "athlete_totals": athlete_totals
    }

    with open(history_file, "w", encoding="utf-8") as f:
        json.dump(history_payload, f, indent=2, ensure_ascii=False)

    print(f"[+] Successfully imported {len(athlete_totals)} verified athletes from {excel_path}!")
    return True


if __name__ == "__main__":
    import_from_excel("aseem.xlsx")
